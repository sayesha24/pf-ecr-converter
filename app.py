"""
PF ECR Converter - Web Interface (Single File Version)
Features:
  1. Convert ECR PDF to Excel (with OCR support for rotated PDFs)
  2. Convert Excel to CSV

Usage:
    1. Install dependencies: pip3 install flask pdfplumber openpyxl pandas pymupdf pytesseract pillow
    2. Install Tesseract OCR: 
       - Mac: brew install tesseract
       - Ubuntu: sudo apt-get install tesseract-ocr
    3. Run: python3 pf_ecr_web_app.py
    4. Open browser: http://localhost:5000
"""

from flask import Flask, request, make_response
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import pandas as pd
import re
import os
import tempfile
from werkzeug.utils import secure_filename

# OCR imports (optional - for rotated PDFs)
try:
    import fitz  # PyMuPDF
    import pytesseract
    from PIL import Image
    import io
    OCR_AVAILABLE = True
except ImportError:
    OCR_AVAILABLE = False

app = Flask(__name__)
app.secret_key = 'pf_ecr_converter_secret_key_2025'

# Health check route
@app.route('/health')
def health():
    return 'OK', 200

# Embedded HTML template
HTML_TEMPLATE = '''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>PF ECR Converter</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body {
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Oxygen, Ubuntu, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            display: flex;
            justify-content: center;
            align-items: center;
            padding: 20px;
        }
        .container {
            background: white;
            border-radius: 20px;
            box-shadow: 0 20px 60px rgba(0, 0, 0, 0.3);
            padding: 40px;
            max-width: 550px;
            width: 100%;
        }
        h1 { color: #333; text-align: center; margin-bottom: 10px; font-size: 24px; }
        .subtitle { color: #666; text-align: center; margin-bottom: 25px; font-size: 14px; }
        
        .tabs {
            display: flex;
            margin-bottom: 25px;
            border-radius: 10px;
            overflow: hidden;
            border: 2px solid #667eea;
        }
        .tab {
            flex: 1;
            padding: 12px 20px;
            text-align: center;
            cursor: pointer;
            background: white;
            color: #667eea;
            font-weight: 600;
            font-size: 14px;
            transition: all 0.3s ease;
            border: none;
        }
        .tab:first-child { border-right: 1px solid #667eea; }
        .tab.active {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
        }
        .tab:hover:not(.active) { background: #f0f2ff; }
        
        .tab-content { display: none; }
        .tab-content.active { display: block; }
        
        .upload-area {
            border: 2px dashed #667eea;
            border-radius: 15px;
            padding: 35px 20px;
            text-align: center;
            cursor: pointer;
            transition: all 0.3s ease;
            background: #f8f9ff;
            margin-bottom: 20px;
        }
        .upload-area:hover { border-color: #764ba2; background: #f0f2ff; }
        .upload-area.dragover { border-color: #764ba2; background: #e8ebff; }
        .upload-icon { font-size: 42px; margin-bottom: 12px; }
        .upload-text { color: #333; font-size: 15px; margin-bottom: 6px; }
        .upload-hint { color: #888; font-size: 12px; }
        input[type="file"] { display: none; }
        
        .file-name {
            background: #e8ebff;
            padding: 12px 20px;
            border-radius: 10px;
            margin-bottom: 20px;
            display: none;
            align-items: center;
            justify-content: space-between;
        }
        .file-name.show { display: flex; }
        .file-name span { color: #333; font-size: 14px; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
        .file-name button { background: none; border: none; color: #e74c3c; cursor: pointer; font-size: 18px; padding: 0 5px; }
        
        .convert-btn {
            width: 100%;
            padding: 14px;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            border: none;
            border-radius: 10px;
            font-size: 15px;
            font-weight: 600;
            cursor: pointer;
            transition: transform 0.2s, box-shadow 0.2s;
        }
        .convert-btn:hover { transform: translateY(-2px); box-shadow: 0 5px 20px rgba(102, 126, 234, 0.4); }
        .convert-btn:disabled { background: #ccc; cursor: not-allowed; transform: none; box-shadow: none; }
        
        .alert { padding: 12px 20px; border-radius: 10px; margin-bottom: 20px; font-size: 14px; }
        .alert-error { background: #ffe8e8; color: #e74c3c; border: 1px solid #f5c6c6; }
        .alert-success { background: #e8ffe8; color: #27ae60; border: 1px solid #c6f5c6; }
        
        .features { margin-top: 25px; padding-top: 20px; border-top: 1px solid #eee; }
        .features h3 { color: #333; font-size: 13px; margin-bottom: 12px; }
        .features ul { list-style: none; color: #666; font-size: 12px; }
        .features li { padding: 4px 0; padding-left: 18px; position: relative; }
        .features li::before { content: "✓"; position: absolute; left: 0; color: #27ae60; font-size: 11px; }
        
        .loading { display: none; text-align: center; padding: 20px; }
        .loading.show { display: block; }
        .spinner {
            border: 3px solid #f3f3f3;
            border-top: 3px solid #667eea;
            border-radius: 50%;
            width: 36px;
            height: 36px;
            animation: spin 1s linear infinite;
            margin: 0 auto 12px;
        }
        @keyframes spin { 0% { transform: rotate(0deg); } 100% { transform: rotate(360deg); } }
    </style>
</head>
<body>
    <div class="container">
        <h1>📄 PF ECR Converter</h1>
        <p class="subtitle">Convert your PF ECR files easily</p>
        
        <div class="tabs">
            <button class="tab active" onclick="switchTab('pdf-to-excel', this)">PDF → Excel</button>
            <button class="tab" onclick="switchTab('excel-to-csv', this)">Excel → CSV</button>
        </div>
        
        {{ERROR_MESSAGE}}

        <div class="tab-content active" id="pdf-to-excel">
            <form method="POST" enctype="multipart/form-data" id="pdfForm">
                <input type="hidden" name="conversion_type" value="pdf_to_excel">
                <div class="upload-area" id="dropZone1">
                    <div class="upload-icon">📑</div>
                    <p class="upload-text">Click to upload or drag & drop</p>
                    <p class="upload-hint">ECR PDF files only</p>
                    <input type="file" name="file" id="fileInput1" accept=".pdf">
                </div>
                <div class="file-name" id="fileName1">
                    <span id="fileNameText1"></span>
                    <button type="button" onclick="clearFile(1)">✕</button>
                </div>
                <div class="loading" id="loading1">
                    <div class="spinner"></div>
                    <p>Converting PDF to Excel...</p>
                </div>
                <button type="submit" class="convert-btn" id="convertBtn1" disabled>Convert to Excel</button>
            </form>
            <div class="features">
                <h3>PDF to Excel features:</h3>
                <ul>
                    <li>Extracts member data from ECR PDF</li>
                    <li>Supports both normal and rotated PDFs</li>
                    <li>Removes name columns (ECR & UAN Repository)</li>
                    <li>Cleans numeric formatting</li>
                </ul>
            </div>
        </div>

        <div class="tab-content" id="excel-to-csv">
            <form method="POST" enctype="multipart/form-data" id="excelForm">
                <input type="hidden" name="conversion_type" value="excel_to_csv">
                <div class="upload-area" id="dropZone2">
                    <div class="upload-icon">📊</div>
                    <p class="upload-text">Click to upload or drag & drop</p>
                    <p class="upload-hint">Excel files (.xlsx, .xls) only</p>
                    <input type="file" name="file" id="fileInput2" accept=".xlsx,.xls">
                </div>
                <div class="file-name" id="fileName2">
                    <span id="fileNameText2"></span>
                    <button type="button" onclick="clearFile(2)">✕</button>
                </div>
                <div class="loading" id="loading2">
                    <div class="spinner"></div>
                    <p>Converting Excel to CSV...</p>
                </div>
                <button type="submit" class="convert-btn" id="convertBtn2" disabled>Convert to CSV</button>
            </form>
            <div class="features">
                <h3>Excel to CSV features:</h3>
                <ul>
                    <li>Converts Excel (.xlsx, .xls) to CSV</li>
                    <li>Removes commas from numbers</li>
                    <li>UTF-8 encoding for proper character support</li>
                    <li>Compatible with any spreadsheet software</li>
                </ul>
            </div>
        </div>
    </div>

    <script>
        function switchTab(tabId, btn) {
            document.querySelectorAll('.tab').forEach(t => t.classList.remove('active'));
            document.querySelectorAll('.tab-content').forEach(c => c.classList.remove('active'));
            document.getElementById(tabId).classList.add('active');
            btn.classList.add('active');
        }

        function setupForm(num) {
            const dropZone = document.getElementById('dropZone' + num);
            const fileInput = document.getElementById('fileInput' + num);
            const fileName = document.getElementById('fileName' + num);
            const fileNameText = document.getElementById('fileNameText' + num);
            const convertBtn = document.getElementById('convertBtn' + num);
            const form = document.getElementById(num === 1 ? 'pdfForm' : 'excelForm');
            const loading = document.getElementById('loading' + num);

            dropZone.addEventListener('click', () => fileInput.click());
            
            fileInput.addEventListener('change', () => {
                if (fileInput.files.length > 0) {
                    fileNameText.textContent = fileInput.files[0].name;
                    fileName.classList.add('show');
                    convertBtn.disabled = false;
                }
            });

            dropZone.addEventListener('dragover', (e) => { e.preventDefault(); dropZone.classList.add('dragover'); });
            dropZone.addEventListener('dragleave', () => { dropZone.classList.remove('dragover'); });
            dropZone.addEventListener('drop', (e) => {
                e.preventDefault();
                dropZone.classList.remove('dragover');
                const files = e.dataTransfer.files;
                if (files.length > 0) {
                    fileInput.files = files;
                    fileNameText.textContent = files[0].name;
                    fileName.classList.add('show');
                    convertBtn.disabled = false;
                }
            });

            form.addEventListener('submit', () => {
                loading.classList.add('show');
                convertBtn.disabled = true;
                dropZone.style.display = 'none';
                fileName.style.display = 'none';
            });
        }

        function clearFile(num) {
            document.getElementById('fileInput' + num).value = '';
            document.getElementById('fileName' + num).classList.remove('show');
            document.getElementById('convertBtn' + num).disabled = true;
        }

        setupForm(1);
        setupForm(2);
    </script>
</body>
</html>'''


def clean_number(value):
    """Remove commas, newlines, and stray characters, then convert to integer if numeric."""
    if value is None:
        return value
    value = str(value).strip()
    
    has_hash_prefix = value.startswith('#')
    if has_hash_prefix:
        value = value[1:].strip()
    
    value = re.sub(r'\n.*$', '', value)
    value = re.sub(r'^[^\d-]+', '', value)
    value = re.sub(r'[^\d]+$', '', value)
    value = value.strip()
    cleaned = value.replace(',', '')
    
    try:
        num_value = int(cleaned)
        if has_hash_prefix:
            return f"# {num_value:,}"
        return num_value
    except ValueError:
        if has_hash_prefix:
            return f"# {value}"
        return value


def extract_ecr_data_pdfplumber(pdf_path):
    """Extract member data from ECR PDF using pdfplumber (for normal PDFs)."""
    all_data = []
    
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            
            if not tables:
                continue
                
            table = tables[0]
            
            for row in table:
                if row and row[0] and str(row[0]).strip().replace('\n', '').isdigit():
                    all_data.append(row)
    
    return all_data


def extract_ecr_data_ocr(pdf_path):
    """Extract member data from ECR PDF using OCR (for rotated/vectorized PDFs)."""
    if not OCR_AVAILABLE:
        return []
    
    all_data = []
    doc = fitz.open(pdf_path)
    
    for page_num in range(len(doc)):
        page = doc[page_num]
        
        # Render page to image at high resolution
        mat = fitz.Matrix(2.5, 2.5)
        pix = page.get_pixmap(matrix=mat)
        img = Image.open(io.BytesIO(pix.tobytes("png")))
        
        # Try rotations to find the right orientation
        for rotation in [0, -90, 90, 180]:
            if rotation != 0:
                img_to_ocr = img.rotate(rotation, expand=True)
            else:
                img_to_ocr = img
            
            # Run OCR
            text = pytesseract.image_to_string(img_to_ocr)
            
            # Find all 12-digit UAN numbers
            uans_found = re.findall(r'\b\d{12}\b', text)
            
            if len(uans_found) >= 3:  # Found meaningful data
                # Parse the text line by line
                lines = text.split('\n')
                
                for line in lines:
                    # Look for UAN pattern (12 digits)
                    uan_match = re.search(r'\b(\d{12})\b', line)
                    if uan_match:
                        uan = uan_match.group(1)
                        
                        # Find serial number before UAN
                        before_uan = line[:uan_match.start()]
                        sl_match = re.search(r'(\d{1,3})\s*[|\s]*$', before_uan)
                        sl_no = sl_match.group(1) if sl_match else str(len(all_data) + 1)
                        
                        # Find all numbers after UAN (wages, contributions, etc.)
                        after_uan = line[uan_match.end():]
                        # Extract numbers (with commas, decimals, or plain)
                        numbers = re.findall(r'[\d,]+\.?\d*', after_uan)
                        
                        # Clean numbers - remove commas
                        clean_numbers = []
                        for num in numbers:
                            cleaned = num.replace(',', '')
                            if cleaned and cleaned.replace('.', '').isdigit():
                                clean_numbers.append(cleaned)
                        
                        # Build row: need at least Gross, EPF, EPS, EDLI, EE, EPS, ER
                        if len(clean_numbers) >= 7:
                            row = [None] * 17
                            row[0] = sl_no  # Sl. No.
                            row[1] = uan    # UAN
                            row[2] = None   # Name ECR (skip)
                            row[3] = None   # Name UAN (skip)
                            
                            # Map numbers to columns
                            # clean_numbers[0] = Gross, [1]=EPF, [2]=EPS, [3]=EDLI, [4]=EE, [5]=EPS, [6]=ER
                            col_map = [4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15]
                            for idx, num in enumerate(clean_numbers[:len(col_map)]):
                                row[col_map[idx]] = num
                            
                            all_data.append(row)
                
                break  # Found data with this rotation
    
    return all_data


def extract_ecr_data(pdf_path):
    """Extract member data from ECR PDF - tries normal extraction first, then OCR."""
    # Try normal extraction first
    data = extract_ecr_data_pdfplumber(pdf_path)
    
    if data:
        return data, "normal"
    
    # If no data found and OCR is available, try OCR
    if OCR_AVAILABLE:
        data = extract_ecr_data_ocr(pdf_path)
        if data:
            return data, "ocr"
    
    return [], None


def convert_ecr_to_excel(pdf_path, output_path):
    """Convert PF ECR PDF to Excel format."""
    
    data, method = extract_ecr_data(pdf_path)
    
    if not data:
        if OCR_AVAILABLE:
            raise ValueError("No member data found in PDF. Please ensure this is a valid ECR PDF.")
        else:
            raise ValueError("No member data found. This PDF may require OCR. Please install: pip install pymupdf pytesseract pillow")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "ECR Data"
    
    # Headers with 2 blank columns after UAN
    headers = [
        'Sl. No.', 'UAN', '', '', 'Gross', 'EPF', 'EPS', 'EDLI', 
        'EE', 'EPS', 'ER', 'NCP\nDays', 'Refunds', 
        'Pension Share', 'ER PF\nShare', 'EE Share', ''
    ]
    
    # Mapping: (pdf_col_index, excel_col_index)
    column_mapping = [
        (0, 1),   # Sl. No.
        (1, 2),   # UAN
        # columns 3 and 4 are blank
        (4, 5),   # Gross
        (5, 6),   # EPF
        (6, 7),   # EPS
        (7, 8),   # EDLI
        (8, 9),   # EE
        (9, 10),  # EPS
        (10, 11), # ER
        (11, 12), # NCP Days
        (12, 13), # Refunds
        (13, 14), # Pension Share
        (14, 15), # ER PF Share
        (15, 16), # EE Share
        (16, 17), # Posting Location
    ]
    
    numeric_columns = [1, 5, 6, 7, 8, 9, 10, 11, 12, 13]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=False)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for row_idx, row_data in enumerate(data, start=2):
        for pdf_col_idx, excel_col_idx in column_mapping:
            if pdf_col_idx < len(row_data) and row_data[pdf_col_idx] is not None:
                value = row_data[pdf_col_idx]
                
                if excel_col_idx in numeric_columns:
                    value = clean_number(value)
                elif value:
                    value = str(value).strip().replace('\n', ' ')
                
                cell = ws.cell(row=row_idx, column=excel_col_idx, value=value)
                cell.alignment = Alignment(horizontal='center' if excel_col_idx <= 2 else 'right')
    
    column_widths = [8, 14, 8, 8, 8, 8, 8, 8, 6, 6, 6, 8, 8, 12, 10, 10, 6]
    for col, width in enumerate(column_widths, start=1):
        col_letter = chr(64 + col) if col <= 26 else 'A' + chr(64 + col - 26)
        ws.column_dimensions[col_letter].width = width
    
    wb.save(output_path)
    return len(data)


def convert_excel_to_csv(excel_path, output_path):
    """Convert Excel file to CSV format with commas removed from numbers."""
    df = pd.read_excel(excel_path, header=None)
    
    if df.empty:
        raise ValueError("The Excel file is empty.")
    
    def remove_commas(val):
        if pd.isna(val):
            return val
        val_str = str(val)
        if ',' in val_str:
            if val_str.startswith('#'):
                prefix = '# '
                num_part = val_str[1:].strip().replace(',', '')
                return prefix + num_part
            else:
                cleaned = val_str.replace(',', '')
                try:
                    float(cleaned)
                    return cleaned
                except ValueError:
                    return val_str
        return val
    
    if hasattr(df, 'map'):
        df = df.map(remove_commas)
    else:
        df = df.applymap(remove_commas)
    
    df.to_csv(output_path, index=False, header=False, encoding='utf-8')
    return len(df)


@app.route('/', methods=['GET', 'POST'])
def index():
    error_message = ''
    
    if request.method == 'POST':
        conversion_type = request.form.get('conversion_type', '')
        
        if 'file' not in request.files:
            error_message = '<div class="alert alert-error">No file selected</div>'
        else:
            file = request.files['file']
            
            if file.filename == '':
                error_message = '<div class="alert alert-error">No file selected</div>'
            else:
                filename = secure_filename(file.filename)
                file_ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''
                
                if conversion_type == 'pdf_to_excel':
                    if file_ext != 'pdf':
                        error_message = '<div class="alert alert-error">Please upload a PDF file for this conversion.</div>'
                    else:
                        with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_pdf:
                            file.save(tmp_pdf.name)
                            pdf_path = tmp_pdf.name
                        
                        base_name = os.path.splitext(filename)[0]
                        excel_filename = f"{base_name}_converted.xlsx"
                        
                        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_xlsx:
                            excel_path = tmp_xlsx.name
                        
                        try:
                            convert_ecr_to_excel(pdf_path, excel_path)
                            
                            with open(excel_path, 'rb') as f:
                                file_data = f.read()
                            
                            os.unlink(pdf_path)
                            os.unlink(excel_path)
                            
                            response = make_response(file_data)
                            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                            response.headers['Content-Disposition'] = f'attachment; filename={excel_filename}'
                            return response
                            
                        except Exception as e:
                            if os.path.exists(pdf_path):
                                os.unlink(pdf_path)
                            if os.path.exists(excel_path):
                                os.unlink(excel_path)
                            error_message = f'<div class="alert alert-error">Error: {str(e)}</div>'
                
                elif conversion_type == 'excel_to_csv':
                    if file_ext not in ['xlsx', 'xls']:
                        error_message = '<div class="alert alert-error">Please upload an Excel file (.xlsx or .xls) for this conversion.</div>'
                    else:
                        with tempfile.NamedTemporaryFile(delete=False, suffix=f'.{file_ext}') as tmp_excel:
                            file.save(tmp_excel.name)
                            excel_path = tmp_excel.name
                        
                        base_name = os.path.splitext(filename)[0]
                        csv_filename = f"{base_name}.csv"
                        
                        with tempfile.NamedTemporaryFile(delete=False, suffix='.csv') as tmp_csv:
                            csv_path = tmp_csv.name
                        
                        try:
                            convert_excel_to_csv(excel_path, csv_path)
                            
                            with open(csv_path, 'rb') as f:
                                file_data = f.read()
                            
                            os.unlink(excel_path)
                            os.unlink(csv_path)
                            
                            response = make_response(file_data)
                            response.headers['Content-Type'] = 'text/csv; charset=utf-8'
                            response.headers['Content-Disposition'] = f'attachment; filename={csv_filename}'
                            return response
                            
                        except Exception as e:
                            if os.path.exists(excel_path):
                                os.unlink(excel_path)
                            if os.path.exists(csv_path):
                                os.unlink(csv_path)
                            error_message = f'<div class="alert alert-error">Error: {str(e)}</div>'
    
    html = HTML_TEMPLATE.replace('{{ERROR_MESSAGE}}', error_message)
    return html


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_DEBUG', 'False').lower() == 'true'
    
    print("\n" + "="*50)
    print("  PF ECR Converter")
    print("="*50)
    print("\n  Features:")
    print("  • PDF to Excel (ECR files)")
    print("  • Excel to CSV")
    print(f"\n  OCR Support: {'✓ Enabled' if OCR_AVAILABLE else '✗ Disabled'}")
    print(f"\n  Open your browser and go to:")
    print(f"  👉  http://localhost:{port}")
    print("\n  Press Ctrl+C to stop the server")
    print("="*50 + "\n")
    app.run(host='0.0.0.0', port=port, debug=debug)
